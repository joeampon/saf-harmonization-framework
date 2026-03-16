"""
data/generate_excel.py
======================
Generate comprehensive Excel workbook with all inputs, outputs, and references.

Sheets
------
1. Literature Database     — 48 studies with full metadata and DOIs
2. Harmonized Values       — raw -> harmonized MFSP and GHG for all studies
3. Parameter Distributions — Monte Carlo parameter ranges with sources
4. Monte Carlo Summary     — P5/P25/median/mean/P75/P95/Std/CV per pathway
5. Sobol Indices           — S1 and ST per parameter per pathway
6. Variance Decomposition  — methodological vs technical variance split
7. External Validation     — real peer-reviewed studies (open-access confirmed), harmonized values
8. Key Findings            — headline numbers from the analysis
9. References              — complete reference list with DOIs

Usage
-----
    from data.generate_excel import create_workbook
    wb = create_workbook(df_harm, mc_results, sobol_results, var_df,
                         ext_df, ext_agg, mc_summary)
    wb.save("outputs/SAF_MetaAnalysis_Harmonization.xlsx")
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               GradientFill)
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_00

import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import PETROLEUM_JET_GHG_WTW, CORSIA_TIER2_THRESHOLD, EU_RED3_THRESHOLD


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _hdr(cell, bg: str = "1F497D", fg: str = "FFFFFF",
          bold: bool = True, sz: int = 10):
    cell.font      = Font(bold=bold, color=fg, size=sz, name="Calibri")
    cell.fill      = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)

def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _data_cell(cell, value=None, fmt=None, bold=False, color=None):
    if value is not None:
        cell.value = value
    if bold:
        cell.font = Font(bold=True, name="Calibri", size=9)
    else:
        cell.font = Font(name="Calibri", size=9)
    if fmt:
        cell.number_format = fmt
    if color:
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = _thin_border()

def _title(ws, text: str, row: int = 1, col: int = 1):
    c = ws.cell(row=row, column=col, value=text)
    c.font      = Font(bold=True, size=13, name="Calibri", color="1F497D")
    c.alignment = Alignment(horizontal="left")

def _set_col_widths(ws, widths: list):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _write_df(ws, df: pd.DataFrame, start_row: int = 3,
              header_bg: str = "1F497D"):
    """Write a DataFrame to a worksheet with styled header."""
    for r_idx, row_data in enumerate(
        dataframe_to_rows(df, index=False, header=True), start=start_row
    ):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == start_row:
                _hdr(cell, bg=header_bg)
            else:
                alt = "F2F2F2" if (r_idx - start_row) % 2 == 0 else "FFFFFF"
                _data_cell(cell, color=alt)


# ---------------------------------------------------------------------------
# Pathway colour map (for per-pathway header colouring)
# ---------------------------------------------------------------------------
PATHWAY_BG = {"ATJ": "1565C0", "HEFA": "2E7D32", "FT-SPK": "E65100", "PtL": "6A1B9A"}


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _sheet_literature(wb: Workbook, df_harm: pd.DataFrame):
    ws = wb.create_sheet("1_Literature_Database")
    _title(ws, "SAF Literature Database — 42 Peer-Reviewed Studies (2009–2022)")
    ws["A2"] = "Harmonization reference: WtWake | Energy allocation | 2023 USD | 10% DR | 90% CF"
    ws["A2"].font = Font(italic=True, size=9, color="555555")

    cols = ["study_id", "authors", "year", "journal", "doi", "pathway",
            "feedstock", "plant_size_tpd", "ref_year_cost", "mfsp_raw",
            "mfsp_unit", "ghg_raw", "allocation", "boundary",
            "include_iluc", "discount_rate", "capacity_factor",
            "plant_lifetime", "currency", "notes"]
    _write_df(ws, df_harm[cols], start_row=3)
    _set_col_widths(ws, [14, 18, 6, 30, 32, 8, 22, 12, 12, 10,
                         8,  8,  16, 9, 10, 12, 14, 12, 8, 40])
    ws.freeze_panes = "A4"


def _sheet_harmonized(wb: Workbook, df_harm: pd.DataFrame):
    ws = wb.create_sheet("2_Harmonized_Values")
    _title(ws, "Harmonized MFSP and GHG — Five-Step Protocol Applied to All 42 Studies")
    ws["A2"] = ("Step1: CPI escalation + unit conversion  |  "
                "Step2: Allocation -> energy basis  |  "
                "Step3: Boundary -> WtWake  |  "
                "Step4: ILUC removal  |  "
                "Step5: CRF normalization (10%/30yr/90%CF)")
    ws["A2"].font = Font(italic=True, size=9, color="555555")

    cols = ["study_id", "authors", "year", "pathway", "feedstock",
            "mfsp_raw", "mfsp_unit", "mfsp_2023_raw", "mfsp_harmonized",
            "ghg_raw", "ghg_harmonized", "ghg_reduction_pct",
            "alloc_correction_factor", "boundary_correction_applied",
            "iluc_removed_gco2e_mj", "crf_correction_factor",
            "capex_fraction_used"]
    _write_df(ws, df_harm[cols], start_row=3)
    _set_col_widths(ws, [14, 18, 6, 8, 22, 10, 8, 14, 14, 10, 14, 12,
                         18, 20, 18, 16, 16])
    ws.freeze_panes = "A4"


def _sheet_params(wb: Workbook):
    ws = wb.create_sheet("3_Parameter_Distributions")
    _title(ws, "Monte Carlo Parameter Distributions — Sources and Justification")

    from data.parameter_distributions import (ATJ_PARAMS, HEFA_PARAMS,
                                               FTSPK_PARAMS, PTL_PARAMS,
                                               METHODOLOGICAL_PARAMS)
    param_sources = {
        # ATJ
        "feedstock_cost":   "USDA ERS corn stover price reports 2015–2023; Tao (2017)",
        "capex_2023":       "Tao (2017) $420M 2011 → CEPCI-escalated; NREL design cases",
        "ethanol_yield":    "Tao (2017) process model; fermentation literature",
        "jet_yield":        "Tao (2017) ATJ conversion step; de Jong (2015)",
        "capacity_factor":  "Biomass industry standard 75–95%; NREL assumption",
        "discount_rate":    "NREL convention 10% Nth-plant; range 5–15% (public/private)",
        "ng_use":           "Tao (2017) energy balance; Han (2017) supplementary",
        "elec_use":         "Tao (2017) process energy analysis",
        "feedstock_ghg":    "GREET 2023 corn stover upstream; regional range",
        "grid_intensity":   "GREET 2023 US avg; EPA eGRID 2023; range covers biogas to coal",
        "alloc_factor":     "ISO 14044 §4.3.4.2; Han (2017) energy/mass/econ. factors",
        "boundary_offset":  "ICAO CORSIA default 3.0 gCO2e/MJ; range 0–5",
        "iluc_penalty":     "Baseline excludes ILUC per harmonization protocol",
        # HEFA
        "h2_use":           "Pearlson (2013); NREL HBio (2020) NP-5100-75060",
        "h2_price":         "IEA (2021) hydrogen supply cost; grey/blue/green range",
        "process_ghg":      "Stratton (2010); Klein (2018) facility LCA",
        # FT
        "ft_efficiency":    "Swanson (2010) NREL TP-6A20-48440; Davis (2018) NREL",
        # PtL
        "elec_cost_mwh":    "IEA WEO 2023; IRENA Renewable Power 2023",
        "elec_capex_kw":    "IEA (2023) Electrolysers; BNEF H2 supply outlook",
        "ft_capex":         "Schmidt (2018); Fasihi (2019) PtL cost breakdown",
        "co2_capture_cost": "IEA (2022) DAC review; IPCC AR6 CCS cost ranges",
    }

    headers = ["Pathway", "Parameter", "Type", "Distribution",
               "Low", "Mode", "High", "Unit", "Source"]
    for c_idx, h in enumerate(headers, start=1):
        _hdr(ws.cell(row=3, column=c_idx), bg="1F497D")
        ws.cell(row=3, column=c_idx).value = h

    row = 4
    for pathway, param_dict in [("ATJ", ATJ_PARAMS), ("HEFA", HEFA_PARAMS),
                                  ("FT-SPK", FTSPK_PARAMS), ("PtL", PTL_PARAMS)]:
        for name, spec in param_dict.items():
            kind = spec[0]
            low  = spec[1] if len(spec) > 1 else ""
            mode = spec[2] if len(spec) > 2 else ""
            high = spec[3] if len(spec) > 3 else ""
            ptype = "Methodological" if name in METHODOLOGICAL_PARAMS else "Technical"
            alt   = "F2F2F2" if (row % 2 == 0) else "FFFFFF"
            vals  = [pathway, name, ptype, kind, low, mode, high, "–",
                     param_sources.get(name, "See data/parameter_distributions.py")]
            for c_idx, v in enumerate(vals, start=1):
                _data_cell(ws.cell(row=row, column=c_idx), value=v, color=alt)
            row += 1

    _set_col_widths(ws, [10, 20, 16, 12, 12, 12, 12, 10, 55])
    ws.freeze_panes = "A4"


def _sheet_mc_summary(wb: Workbook, mc_summary: pd.DataFrame):
    ws = wb.create_sheet("4_Monte_Carlo_Summary")
    _title(ws, "Monte Carlo Results — 10,000 Iterations per Pathway")
    ws["A2"] = "Seed=42; parameters sampled from distributions in Sheet 3"
    ws["A2"].font = Font(italic=True, size=9, color="555555")
    _write_df(ws, mc_summary, start_row=3)
    _set_col_widths(ws, [10, 20, 10, 10, 10, 10, 10, 10, 10, 10])
    ws.freeze_panes = "A4"


def _sheet_sobol(wb: Workbook, sobol_df: pd.DataFrame):
    ws = wb.create_sheet("5_Sobol_Indices")
    _title(ws, "Sobol First-Order (S1) and Total-Order (ST) Sensitivity Indices")
    ws["A2"] = ("Jansen (1999) estimator; N=1500 base samples; "
                "Indices <0.05 flagged as negligible (high sampling uncertainty)")
    ws["A2"].font = Font(italic=True, size=9, color="555555")
    _write_df(ws, sobol_df, start_row=3)
    _set_col_widths(ws, [10, 22, 16, 10, 10, 10, 10, 16, 14])
    ws.freeze_panes = "A4"


def _sheet_variance(wb: Workbook, var_df: pd.DataFrame):
    ws = wb.create_sheet("6_Variance_Decomposition")
    _title(ws, "Variance Decomposition — Methodological vs. Technical/Economic")
    ws["A2"] = "CV after = CV before × sqrt(1 - S1_methodological_fraction)"
    ws["A2"].font = Font(italic=True, size=9, color="555555")
    _write_df(ws, var_df, start_row=3)
    _set_col_widths(ws, [10, 8, 18, 14, 14, 14, 22])
    ws.freeze_panes = "A4"


def _sheet_ext_validation(wb: Workbook, ext_df: pd.DataFrame, ext_agg: pd.DataFrame):
    ws = wb.create_sheet("7_External_Validation")
    _title(ws, "External Validation — Real Peer-Reviewed Studies (open-access, full text confirmed)")
    ws["A2"] = ("Only studies where ALL required values (MFSP, GHG, allocation, boundary, DR, CF) "
                "were confirmed from open-access full text are included. "
                "See analysis/variance_decomposition.py for commented stubs of additional papers "
                "pending full-text retrieval. CV reduction is computed only for pathways with ≥ 2 entries.")
    ws["A2"].font = Font(italic=True, size=9, color="555555")
    _write_df(ws, ext_df, start_row=3)
    ws["A20"] = "Aggregate CV statistics by pathway"
    ws["A20"].font = Font(bold=True, size=10, color="1F497D")
    _write_df(ws, ext_agg, start_row=21)
    _set_col_widths(ws, [16, 18, 6, 10, 14, 16, 14, 16, 16, 14])


def _sheet_key_findings(wb: Workbook, df_harm: pd.DataFrame,
                         mc_results: dict, var_df: pd.DataFrame):
    ws = wb.create_sheet("8_Key_Findings")
    _title(ws, "Key Findings — SAF Harmonization Meta-Analysis v2.0")

    findings = [
        ["Category", "Finding", "Value", "Notes"],
        ["Dataset", "Total studies analyzed", 42, "ATJ=12, HEFA=11, FT-SPK=11, PtL=8"],
        ["Dataset", "Publication year range", "2009–2022", ""],
        ["Dataset", "Pathways covered", 4, "ATJ, HEFA, FT-SPK, PtL"],
        ["Cost",    "Raw MFSP range (2023 $/GGE)",
         f"${df_harm['mfsp_2023_raw'].min():.2f}–${df_harm['mfsp_2023_raw'].max():.2f}",
         "Before harmonization"],
        ["Cost",    "Harmonized MFSP range (2023 $/GGE)",
         f"${df_harm['mfsp_harmonized'].min():.2f}–${df_harm['mfsp_harmonized'].max():.2f}",
         "After harmonization to common basis"],
        ["GHG",     "Raw GHG range (gCO2e/MJ)",
         f"{df_harm['ghg_raw'].min():.1f}–{df_harm['ghg_raw'].max():.1f}",
         "Before harmonization"],
        ["GHG",     "Harmonized GHG range (gCO2e/MJ)",
         f"{df_harm['ghg_harmonized'].min():.1f}–{df_harm['ghg_harmonized'].max():.1f}",
         "After harmonization"],
        ["GHG",     "Petroleum jet baseline (gCO2e/MJ)", PETROLEUM_JET_GHG_WTW,
         "ICAO CORSIA WtWake (Doc 10164, 2022)"],
        ["GHG",     "Avg GHG reduction vs petroleum",
         f"{df_harm['ghg_reduction_pct'].mean():.0f}%",
         "Mean across all 42 harmonized studies"],
        ["Framework", "CV reduction (MFSP)",
         f"{var_df[var_df.Metric=='MFSP']['Variance_Reduction_pct'].mean():.0f}% avg",
         "Across 4 pathways"],
        ["Framework", "CV reduction (GHG)",
         f"{var_df[var_df.Metric=='GHG']['Variance_Reduction_pct'].mean():.0f}% avg",
         "Across 4 pathways"],
    ]

    for pathway in ["ATJ", "HEFA", "FT-SPK", "PtL"]:
        for col, metric in [("mfsp", "MFSP ($/GGE)"), ("ghg", "GHG (gCO2e/MJ)")]:
            mc   = mc_results[pathway][col]
            med  = mc.median()
            p5   = np.percentile(mc, 5)
            p95  = np.percentile(mc, 95)
            findings.append([
                pathway,
                f"MC {metric} [P5–P50–P95]",
                f"${p5:.2f}–${med:.2f}–${p95:.2f}"
                if col == "mfsp" else f"{p5:.1f}–{med:.1f}–{p95:.1f}",
                "10,000 Monte Carlo iterations",
            ])

    for r_idx, row_data in enumerate(findings, start=3):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 3:
                _hdr(cell, bg="1F497D")
            else:
                alt = "EBF5EB" if row_data[0] in ("ATJ","HEFA","FT-SPK","PtL") else "FFFFFF"
                _data_cell(cell, color=alt)

    _set_col_widths(ws, [14, 38, 28, 45])


def _sheet_references(wb: Workbook):
    ws = wb.create_sheet("9_References")
    _title(ws, "Complete Reference List — SAF Harmonization Meta-Analysis")

    refs = [
        ["#", "Citation", "DOI", "Role in Analysis"],
        # Primary database
        [1,  "Tao et al. (2017) Green Chemistry",
         "10.1039/C6GC02800D", "ATJ TEA baseline (TAO2017)"],
        [2,  "Han et al. (2017) Biotechnology for Biofuels",
         "10.1186/s13068-017-0698-z", "ATJ LCA baseline (HAN2017)"],
        [3,  "Yao et al. (2017) Biotechnology for Biofuels",
         "10.1186/s13068-017-0702-7", "ATJ TEA (YAO2017)"],
        [4,  "Zhao et al. (2021) Applied Energy",
         "10.1016/j.apenergy.2021.116784", "ATJ TEA (ZHAO2021)"],
        [5,  "Capaz et al. (2021) Resources, Conservation and Recycling",
         "10.1016/j.resconrec.2020.105260", "ATJ LCA (CAPAZ2021)"],
        [6,  "de Jong et al. (2017) Biofuels Bioprod. Biorefining",
         "10.1002/bbb.1745", "ATJ TEA EUR (JONG2017)"],
        [7,  "Moreira et al. (2014) Bioresource Technology",
         "10.1016/j.biortech.2014.06.099", "ATJ TEA Brazil (MOREIRA2014)"],
        [8,  "Wang et al. (2019) Green Chemistry",
         "10.1039/C9GC01747G", "ATJ MSW TEA (WANG2019A)"],
        [9,  "Zhang et al. (2020) Energy & Environmental Science",
         "10.1039/D0EE01435G", "ATJ TEA/LCA (ZHANG2020)"],
        [10, "Michailos et al. (2019) Sustainable Energy & Fuels",
         "10.1039/C8SE00487K", "ATJ TEA UK (MICHAILOS2019)"],
        [11, "Bann et al. (2017) Bioresource Technology",
         "10.1016/j.biortech.2017.06.141", "ATJ harmonized stochastic (BANN2017)"],
        [12, "Hari et al. (2015) Renewable and Sustainable Energy Reviews",
         "10.1016/j.rser.2014.10.095", "ATJ TEA (HARI2015)"],
        [13, "Pearlson et al. (2013) Biofuels Bioprod. Biorefining",
         "10.1002/bbb.1414", "HEFA TEA (PEARLSON2013)"],
        [14, "Wong et al. (2013) Energy Policy",
         "10.1016/j.enpol.2013.07.106", "HEFA TEA Jatropha (WONG2013)"],
        [15, "Shonnard et al. (2010) Environmental Science & Technology",
         "10.1021/es103085m", "HEFA LCA Camelina (SHONNARD2010)"],
        [16, "Staples et al. (2014) Energy & Environmental Science",
         "10.1039/C3EE44096A", "HEFA TEA/LCA (STAPLES2014)"],
        [17, "Klein et al. (2018) Journal of Cleaner Production",
         "10.1016/j.jclepro.2018.03.161", "HEFA UCO TEA (KLEIN2018)"],
        [18, "Stratton et al. (2010) MIT LAE Report",
         "10.21949/1503647", "HEFA Palm LCA + ILUC (STRATTON2010)"],
        [19, "Ramos et al. (2019) Fuel",
         "10.1016/j.fuel.2019.05.078", "HEFA Algae TEA (RAMOS2019)"],
        [20, "Tanzil et al. (2021) Biomass and Bioenergy",
         "10.1016/j.biombioe.2021.106098", "HEFA Carinata (TANZIL2021)"],
        [21, "Pavlenko et al. (2019) ICCT Working Paper",
         "10.13140/RG.2.2.32762.31682", "HEFA Soy + ILUC (PAVLENKO2019)"],
        [22, "Hanif et al. (2021) Energy Conversion and Management",
         "10.1016/j.enconman.2021.114167", "HEFA WCO (HANIF2021)"],
        [23, "Geleynse et al. (2018) ChemSusChem",
         "10.1002/cssc.201801581", "HEFA Camelina (GELEYNSE2018)"],
        [24, "Swanson et al. (2010) NREL TP-6A20-48440",
         "10.2172/1007688", "FT-SPK NREL baseline (SWANSON2010)"],
        [25, "Diederichs et al. (2016) Energy",
         "10.1016/j.energy.2016.05.080", "FT-SPK Bagasse (DIEDERICHS2016)"],
        [26, "Liu et al. (2013) Renewable Energy",
         "10.1016/j.renene.2013.01.047", "FT-SPK Forestry (LIU2013)"],
        [27, "Trippe et al. (2013) Fuel Processing Technology",
         "10.1016/j.fuproc.2013.06.024", "FT-SPK Straw EUR (TRIPPE2013)"],
        [28, "Larson et al. (2009) Biofuels Bioprod. Biorefining",
         "10.1002/bbb.137", "FT-SPK CCS negative GHG (LARSON2009)"],
        [29, "Thakkar et al. (2019) Biofuels",
         "10.1080/17597269.2019.1660163", "FT-SPK Poplar (THAKKAR2019)"],
        [30, "Susmozas et al. (2014) Int. J. Hydrogen Energy",
         "10.1016/j.ijhydene.2014.03.196", "FT-SPK Eucalyptus EUR (SUSMOZAS2014)"],
        [31, "Dimitriou et al. (2018) Energy & Environmental Science",
         "10.1039/C7EE02819A", "FT-SPK MSW (DIMITRIOU2018)"],
        [32, "Lane et al. (2021) Int. J. Life Cycle Assessment",
         "10.1007/s11367-021-01956-y", "FT-SPK Woody biomass (LANE2021)"],
        [33, "Marcucci et al. (2022) Applied Energy",
         "10.1016/j.apenergy.2022.118654", "FT-SPK Agri residues (MARCUCCI2022)"],
        [34, "Hillestad et al. (2018) Fuel",
         "10.1016/j.fuel.2018.06.048", "FT-SPK Forestry (HILLESTAD2018)"],
        [35, "Schmidt et al. (2018) Joule",
         "10.1016/j.joule.2018.05.008", "PtL foundational (SCHMIDT2018)"],
        [36, "Brynolf et al. (2018) Renewable Sustainable Energy Reviews",
         "10.1016/j.rser.2017.05.183", "PtL wind (BRYNOLF2018)"],
        [37, "Hombach et al. (2019) Journal of Cleaner Production",
         "10.1016/j.jclepro.2019.03.310", "PtL wind+DAC (HOMBACH2019)"],
        [38, "Fasihi et al. (2019) Joule",
         "10.1016/j.joule.2019.05.002", "PtL solar MENA (FASIHI2019)"],
        [39, "Ueckerdt et al. (2021) Nature Climate Change",
         "10.1038/s41558-021-01032-7", "PtL global potential (UECKERDT2021)"],
        [40, "Tremel et al. (2015) Int. J. Hydrogen Energy",
         "10.1016/j.ijhydene.2015.05.011", "PtL small scale (TREMEL2015)"],
        [41, "Becattini et al. (2021) Joule",
         "10.1016/j.joule.2021.01.013", "PtL solar+DAC 2050 (BECATTINI2021)"],
        [42, "Niermann et al. (2021) Energy Conversion and Management",
         "10.1016/j.enconman.2021.113743", "PtL carrier compare (NIERMANN2021)"],
        # Methodology references
        ["M1", "ISO 14044:2006 Environmental management — LCA",
         "ISO standard", "Allocation methodology §4.3.4.2"],
        ["M2", "ICAO CORSIA Doc 10164 (2022)",
         "ICAO official", "GHG baseline 89 gCO2e/MJ; WtG->WtWake delta"],
        ["M3", "Davis et al. (2018) NREL/TP-5100-71949",
         "10.2172/1493009", "TEA methodology; CRF; NREL convention"],
        ["M4", "GREET 2023 (Wang et al.)",
         "10.2172/1824336", "Background emission factors; grid intensity"],
        ["M5", "Jansen (1999) Computer Physics Communications",
         "10.1016/S0010-4655(98)00154-4", "Sobol estimator method"],
        ["M6", "Searchinger et al. (2008) Science",
         "10.1126/science.1151861", "ILUC methodology basis"],
        ["M7", "IEA (2023) Electrolysers Technology Roadmap",
         "IEA report", "PtL electrolyzer cost range"],
        ["M8", "IEA World Energy Outlook 2023",
         "IEA report", "Electricity cost range for PtL"],
        ["M9", "IPCC AR6 WG3 (2022) Chapter 7",
         "10.1017/9781009157926.009", "Renewable electricity GHG; ILUC ranges"],
        ["M10","Chemical Engineering CEPCI 2005-2023",
         "Chemical Engineering Magazine", "Capital cost escalation index"],
        ["M11","US BLS CPI-U All Items 2005-2023",
         "www.bls.gov/cpi", "Operating cost escalation"],
        ["M12","US Federal Reserve H.10 EUR/USD rates",
         "www.federalreserve.gov", "Currency conversion"],
        ["M13","Prussi et al. (2021) Renewable & Sustainable Energy Reviews",
         "10.1016/j.rser.2021.111614",
         "Land use / water use / feedstock availability scores (Fig K radar)"],
        ["M14","ICAO CORSIA Eligible Fuels Life Cycle Assessment (2022)",
         "ICAO official", "TRL basis and land/water qualitative scores (Fig K radar)"],
    ]

    for r_idx, row_data in enumerate(refs, start=3):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 3:
                _hdr(cell)
            else:
                alt = "F0F4FF" if str(row_data[0]).startswith("M") else "FFFFFF"
                _data_cell(cell, color=alt)
                cell.alignment = Alignment(horizontal="left", wrap_text=True)

    _set_col_widths(ws, [6, 55, 30, 40])
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def create_workbook(df_harm, mc_results, sobol_df, var_df,
                    ext_df, ext_agg, mc_summary) -> Workbook:
    """
    Build complete Excel workbook with all analysis outputs.

    Parameters
    ----------
    df_harm     : pd.DataFrame  harmonized literature dataset
    mc_results  : dict          Monte Carlo results per pathway
    sobol_df    : pd.DataFrame  Sobol indices tidy DataFrame
    var_df      : pd.DataFrame  variance decomposition
    ext_df      : pd.DataFrame  external validation study-level
    ext_agg     : pd.DataFrame  external validation aggregate
    mc_summary  : pd.DataFrame  Monte Carlo summary statistics

    Returns
    -------
    openpyxl.Workbook
    """
    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    _sheet_literature(wb, df_harm)
    _sheet_harmonized(wb, df_harm)
    _sheet_params(wb)
    _sheet_mc_summary(wb, mc_summary)
    _sheet_sobol(wb, sobol_df)
    _sheet_variance(wb, var_df)
    _sheet_ext_validation(wb, ext_df, ext_agg)
    _sheet_key_findings(wb, df_harm, mc_results, var_df)
    _sheet_references(wb)

    return wb
